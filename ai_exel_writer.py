import openpyxl
from pathlib import Path
import string
from openpyxl.styles import (
    Alignment, Font)


def universal_xlsx_writer(columns_names, row_data, file_path, sheet_name):
    def set_column_widths():
        alphabet = string.ascii_uppercase
        for i in range(len(columns_names)):
            ws.column_dimensions[alphabet[i]].width = 20
            ws.cell(row=1, column=i + 1).font = Font(color="FF0000", size=14, bold=True)
            ws.cell(row=1, column=i + 1).alignment = Alignment(horizontal='center')

    # Create the file if it doesn't exist yet
    if not Path(file_path).is_file():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        set_column_widths()
x
        wb.save(file_path)

    wb = openpyxl.load_workbook(file_path, read_only=False)
    ws = wb[sheet_name]

    # Write the column headers if sheet is empty
    if ws.max_row == 1:
        for col_num, column in enumerate(columns_names, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = column

    # Write the row data
    last_line = ws.max_row
    for col_num, cell_value in enumerate(row_data, 1):
        cell = ws.cell(row=last_line + 1, column=col_num)
        ws.cell(row=last_line + 1, column=col_num).alignment = Alignment(wrap_text=True, horizontal='left')
        cell.value = cell_value

    wb.save(file_path)


if __name__ == '__main__':
    columns = ['Name', 'Age', 'Job', 'F0x']
    row_data = ['mouse', 'John', 30, 'Analyst']
    file_path = '/Users/evgeniy/Documents/test33.xlsx'
    sheet_name = 'Sheet_name'

    universal_xlsx_writer(columns, row_data, file_path, sheet_name)
    # create_excel(columns, row_data, file_path, sheet_name)
