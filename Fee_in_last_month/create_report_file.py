from openpyxl import load_workbook, Workbook
import os
from pathlib import Path
from typing import Optional


def create_report_file(month_name: str, html_folder: str, photographer: Optional[str] = None):
    report_folder = Path(html_folder).parent
    path_to_file = report_folder / f"report_file_{month_name}.xlsx"

    def set_column_widths():
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25  # задаю ширину колонки
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 100

    if os.path.exists(path_to_file):
        wb = load_workbook(filename=path_to_file, read_only=False)
        # IDEA  - test sheet is it exist don't create double  !!
        if photographer not in wb.sheetnames:
            ws = wb.create_sheet(photographer)
        ws = wb[photographer]
    else:
        wb = Workbook()  # если файла еще нет
        ws = wb.active  # если файла еще нет
        ws.title = photographer  # если файла еще нет

    set_column_widths()

    headers = {
        'A1': 'number',
        'B1': 'KSP_id',
        'C1': 'date of publication',
        'D1': 'publication',
        'E1': 'material',
    }

    for cell, value in headers.items():
        ws[cell] = value

    wb.save(path_to_file)
    return path_to_file
