import openpyxl
from pathlib import Path
import string
from openpyxl.styles import (
    Alignment, Font)


#  set columns widhth in sheet header
def set_column_widths(ws, columns_names, column_width=20):
    alphabet = string.ascii_uppercase
    for i in range(len(columns_names)):
        ws.column_dimensions[alphabet[i]].width = column_width
        ws.cell(row=1, column=i + 1).font = Font(color="FF0000", size=14, bold=True)
        ws.cell(row=1, column=i + 1).alignment = Alignment(horizontal='center')


# Write the column headers if sheet is empty
def write_column_headers(ws, columns_names):
    if ws.max_row == 1:
        for col_num, column in enumerate(columns_names, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = column


# Write  data from list in  the row
def write_list_in_line(ws, row_data):
    last_line = ws.max_row
    for col_num, cell_value in enumerate(row_data, 1):
        cell = ws.cell(row=last_line + 1, column=col_num)
        ws.cell(row=last_line + 1, column=col_num).alignment = Alignment(wrap_text=True, horizontal='center')
        cell.value = cell_value


# Write  data from list in  the column
def write_list_to_column(ws, column_data):
    last_column = ws.max_column
    for row_num, cell_value in enumerate(column_data, 1):
        cell = ws.cell(row=row_num, column=last_column + 1)
        ws.cell(row=row_num, column=last_column + 1).alignment = Alignment(wrap_text=True, horizontal='center')
        cell.value = cell_value


def write_to_cell(ws, row_line, column_number, cell_data, photographer):
    cell = ws.cell(row=row_line + 1, column=1)
    cell.value = photographer
    cell = ws.cell(row=row_line + 1, column=column_number + 1)
    ws.cell(row=row_line + 1, column=column_number + 1).alignment = Alignment(wrap_text=True, horizontal='center')
    cell.value = cell_data


def universal_xlsx_writer(photographer, columns_names, file_path, sheet_name, row_line, column_number, cell_data ):
    # Create the file if it doesn't exist yet
    if not Path(file_path).is_file():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        set_column_widths(ws, columns_names)
        wb.save(file_path)

    wb = openpyxl.load_workbook(file_path, read_only=False)

    # check sheetname and create new if it doesn't exist
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.title = sheet_name
        set_column_widths(ws, columns_names)
        write_column_headers(ws, columns_names)
    ws = wb[sheet_name]
    write_column_headers(ws, columns_names)

    write_to_cell(ws, row_line, column_number, cell_data, photographer)

    wb.save(file_path)


if __name__ == '__main__':
    columns_n = (
        'Name', 'January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October',
        'November', 'December')
    # row_data_t = 'column 11 date'
    list_data_t = [44, 74444, 33, 7777]  # only for writing by rows
    sheet_name_t = '2'
    column_number_t = 2
    cell_data_t = 'gjksgeggjw'

    universal_xlsx_writer(photographer='Pupkin',
                          columns_names=columns_n,
                          file_path='/Users/evgeniy/Documents/test33.xlsx',
                          sheet_name=sheet_name_t,
                          row_line=7,
                          column_number=7,
                          cell_data='7x7'

                          )
    # def universal_xlsx_writer(columns_names, file_path, sheet_name, row_line, column_number, cell_data):
