import openpyxl
from openpyxl.styles import (
    Alignment, Font
)


def write_to_xlsx(image_id, action, shoot_id, count, path_to_file):
    wb = openpyxl.load_workbook(path_to_file, read_only=False)

    sheet = wb[shoot_id]

    sheet.cell(row=3 + count, column=1).alignment = Alignment(horizontal='center')
    sheet.cell(row=3 + count, column=1).font = Font(size=12, bold=True)
    sheet.cell(row=3 + count, column=1).value = image_id
    sheet.cell(row=3 + count, column=2).alignment = Alignment(wrap_text=True, horizontal='center')
    sheet.cell(row=3 + count, column=2).value = action

    wb.save(path_to_file)


if __name__ == '__main__':
    write_to_xlsx('KSP_016152_555', 'deleted', 'KSP_016152', 3,
                  '/Volumes/big4photo/Documents/Kommersant/deleted_images.xlsx')
